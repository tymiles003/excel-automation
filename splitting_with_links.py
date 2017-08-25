#python script for splitting the master consolidated forecast file on the basis of combination column- 'ClientEx_ProcessVP' i.e. Client Executive & Process VP
#Developed by- Pankaj Kumar
#Libraries required- Pandas, numpy, xlsxwriter, openpyxl, xlrd, xlwt
#Required input/master file- "Consolidated Forecast.xlsx"
#UPDATES-
# Date          Updated by              Decription
#30/07/2017     Pankaj Kumar            15 sheets including Summary and Dashboard
#16/08/2017     Pankaj                  Blank columns hidden, Rename blank col1 as LoB in rev_f_fin/QoQ/MoM 
#17/08/2017     Pankaj                  Added a column sort key for sorting on the basis of client name and location

#-------------------------------------------------------part 1: splitting the data on the basis of key column-----------------------------------------------------
#1 importing libraries
import pandas as pd
from openpyxl import load_workbook
import os
from xlsxwriter.utility import xl_rowcol_to_cell

#define the location of the master file along with file name
path = r'C:\Users\pankaj111056\Desktop\FP&A\Automation_FP&A\Version1.0\Test\Splitting\Consolidated Forecast.xlsx'
#set the file as workbook
wb = load_workbook(path,data_only=False)
sheet_list=wb.get_sheet_names()

#2 sheets of the workbook along with links
modify_raw=wb.get_sheet_by_name(sheet_list[0])#Sheet1- Modify Data
fte_raw=wb.get_sheet_by_name(sheet_list[1])#Sheet2- FTE Forecast-2017
rev_f_com_raw=wb.get_sheet_by_name(sheet_list[2])#sheet3- Rev Forecast Committed
pass_rev_raw=wb.get_sheet_by_name(sheet_list[3])#sheet4- Passthrough Revenue
opp_inc_raw=wb.get_sheet_by_name(sheet_list[4])#sheet5- Opportunities Included
rev_f_fin_raw=wb.get_sheet_by_name(sheet_list[5])#sheet6- Revenue Forecast Final
qoq_raw=wb.get_sheet_by_name(sheet_list[6])#sheet7- QoQ Details
mom_raw=wb.get_sheet_by_name(sheet_list[7])#sheet8- MoM Details with Location
summary_raw=wb.get_sheet_by_name(sheet_list[8])#sheet9- Summary
hist_data=wb.get_sheet_by_name(sheet_list[9])#sheet10- Historic Data
raw_data=wb.get_sheet_by_name(sheet_list[10])#sheet11- Raw Data
dashboard=wb.get_sheet_by_name(sheet_list[11])#sheet12- Dashboard
fx=wb.get_sheet_by_name(sheet_list[12])#sheet13- Fx
fte18_raw=wb.get_sheet_by_name(sheet_list[13])#sheet14- FTE Forecast- 2018
rev18_raw=wb.get_sheet_by_name(sheet_list[14])#sheet15- Rev Forecast- 2018

#3 changing openpyxl sheets into pandas dataframes
#sheet1- Modify Data
modify = pd.DataFrame(modify_raw.values, columns=['Client Code', 'Client Name', 'Client Executive', 'Ops Lead', 'Process VP', 'ClientEx_ProcessVP', 'Vertical',
    'Location','Client Status', 'Modify Data (Yes/No)','Blank Col1','Blank Col2','Blank Col3','Key','Final Key','Sort Key'])
modify=modify[4:]

#sheet2- FTE Forecast- 2017
fte = pd.DataFrame(fte_raw.values, columns=['Client Code','Client Name','Client Executive', 'Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
    'Client Status', 'Blank Col1','Blank Col2', 'Blank Col3','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17',
    'Dec 17','FY 17','','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','','Jan 17', 'Feb 17',
    'Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
    'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Key','Sort Key'])
fte=fte[4:]

#sheet3- Rev Forecast Committed
rev_f_com = pd.DataFrame(rev_f_com_raw.values, columns=['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical',
    'Location','Client Status','Blank col1','Blank col2','Blank col3','Billing Rate','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17',
    'Sep 17','Oct 17','Nov 17','Dec 17','FY 17','','Q1 17','Q2 17','Q3 17','Q4 17','FY 17',' ', 'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17',
    'Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
    'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total','','Reasons for Variance against Previous forecast',
    'Key','Sort Key'])
rev_f_com=rev_f_com[4:]

#sheet4- Passthrough Revenue
pass_rev = pd.DataFrame(pass_rev_raw.values, columns=['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical',
    'Location', 'Client Status', 'Blank col1','Blank col2','Blank col3','Billing Rate','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17',
    'Sep 17','Oct 17','Nov 17','Dec 17','FY 17','','Q1 17','Q2 17','Q3 17','Q4 17','FY 17',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17',
    'Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17',
    'Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total','','Reasons for Variance against Previous forecast','','Jan 17', 'Feb 17',
    'Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total',' ','Jan 17',
    'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total','',
    'Reasons for Variance against Budget','Key','Sort Key'])
pass_rev=pass_rev[4:]

#sheet5- Opportunities Included
opp_inc = pd.DataFrame(opp_inc_raw.values, columns=['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
    'Client Status','Blank col1','Blank col2','Blank col3','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17',
    'Dec 17','Total',' ', 'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17', 'Total', ' ','Q1 17',
    'Q2 17','Q3 17','Q4 17','Total', ' ', 'Probability', 'Committed - Unsigned', 'Yet to be won*', 'Total', 'Check', ' ','Jan 17', 'Feb 17','Mar 17','Apr 17',
    'May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total',' ','Q1 17','Q2 17','Q3 17','Q4 17','Total', ' ','Jan 17', 'Feb 17','Mar 17',
    'Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total',' ','Q1 17','Q2 17','Q3 17','Q4 17','Total', ' ','Jan 17', 'Feb 17',
    'Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total',' ','Committed Unsigned details', 'Yet-to-be won details',
    'Key','Final Key','Sort Key'])
opp_inc=opp_inc[4:]

#sheet6- Revenue Forecast Final
rev_f_fin = pd.DataFrame(rev_f_fin_raw.values, columns=['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical',
    'Location','Client Status','LoB','Blank col2','Blank col3','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17',
    'Nov 17','Dec 17','Total',' ','Q1 17','Q2 17','Q3 17','Q4 17','Total', ' ','Committed Signed', 'Committed Unsigned', 'Yet to be won*', 'Total', 'Check',
    'Q1 PF', 'var','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17',' ','Q1 17','Q2 17',
    'Q3 17','Q4 17','FY 17',' ', 'Committed Signed', 'Committed Unsigned', 'Yet to be won*', 'FY 17', 'Check',' ', 'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17',
    'Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17',' ','Q1 17','Q2 17','Q3 17','Q4 17','Total',' ','Committed Signed', 'Committed Unsigned',
    'Yet to be won*', 'FY 17',' ','COMMENTS ON VARIANCE AGAINST BUDGET', 'Inputs received','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17',
    'Sep 17','Oct 17','Nov 17','Dec 17','FY 17',' ','Committed Signed', 'Committed Unsigned', 'Yet to be won*', 'FY 17',' ','Jan 17', 'Feb 17','Mar 17','Apr 17',
    'May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17',' ','Committed Signed', 'Committed Unsigned', 'Yet to be won*', 'FY 17',' ',
    'COMMENTS ON VARIANCE(CF-PF)', 'Q1 17 variance', 'Q2 17 variance', 'Q3 17 variance', 'Q4 17 variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
    'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17',
    'Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17',
    'FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ',
    'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17',
    'Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
    'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17',
    'Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17',
    'FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance','Key'
    ,'Sort Key'])
rev_f_fin=rev_f_fin[4:]

#sheet7- QoQ Details
qoq = pd.DataFrame(qoq_raw.values, columns=['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
    'Client Status','LoB','Blank col2','Blank col3','Committed Signed', 'Committed - Unsigned','Yet to be won*','NCR','TOTAL','','Committed Signed',
    'Committed - Unsigned','Yet to be won*','NCR','TOTAL','','Committed Signed', 'Committed - Unsigned','Yet to be won*','NCR','TOTAL','',
    'Committed Signed', 'Committed - Unsigned','Yet to be won*','NCR','TOTAL','','Committed Signed', 'Committed - Unsigned','Yet to be won*','NCR','TOTAL','Key',
    'Final Key','Sort Key'])

#sheet8- MoM Details with Location
mom = pd.DataFrame(mom_raw.values, columns=['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
    'Client Status','LoB','Blank col2','Blank col3','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
    'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
    'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
    'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
    'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
    'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
    'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','Key','Sort Key'])

#sheet9- Summary
summary= pd.DataFrame(summary_raw.values)
#sheet10- Historic Data
hist_data= pd.DataFrame(hist_data.values)
#sheet11- Raw Data
raw_data= pd.DataFrame(raw_data.values)
#sheet12- Dashboard
dashboard= pd.DataFrame(dashboard.values)
#sheet13- Fx
fx= pd.DataFrame(fx.values)


#sheet14- FTE Forecast- 2018
fte18 = pd.DataFrame(fte18_raw.values, columns=['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
    'Client Status','Blank col1','Blank col2','Blank col3','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18',
    'Dec 18','FY 18','','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','',
    'Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','',
    'Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','',
    'Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','Key','Sort Key'])

#sheet15- Rev Forecast- 2018
rev18 = pd.DataFrame(rev18_raw.values, columns=['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
    'Client Status','Blank col1','Blank col2','Blank col3','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18',
    'Dec 18','FY 18','','Q1 18','Q2 18','Q3 18','Q4 18','FY 18',' ','Probability %','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18',
    'Sep 18','Oct 18','Nov 18','Dec 18','FY 18','','Q1 18','Q2 18','Q3 18','Q4 18','FY 18',' ','Probability %','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18',
    'Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','','Q1 18','Q2 18','Q3 18','Q4 18','FY 18',' ','Jan 18', 'Feb 18','Mar 18','Apr 18',
    'May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','','Q1 18','Q2 18','Q3 18','Q4 18','FY 18',' ','Comments','Key','Sort Key',])

#4 generating unique combination of ClientEx_ProcessVP for naming the output files & splitting on this basis
modify.set_index(keys=['ClientEx_ProcessVP'], drop=False,inplace=True)
key=modify['ClientEx_ProcessVP'].unique().tolist()
count=len(key)

#5 naming the file as per the unique combinations and splitting the data into different dataframes then writing back on each excel file all 11 sheets
for i in range(0,count):
    #renaming files as key
    x=key[i]+'.xlsx'
    writer = pd.ExcelWriter(x)
    #sheet1 for ith key
    df1=modify.loc[modify.ClientEx_ProcessVP==key[i]]
    df1.to_excel(writer,'Modify Data',header=None, index=False)
    #sheet2
    df2=fte.loc[fte.ClientEx_ProcessVP==key[i]]
    df2.to_excel(writer,'FTE Forecast- 2017',header=None,index=False)
    #sheet3
    df3=rev_f_com.loc[rev_f_com.ClientEx_ProcessVP==key[i]]
    df3.to_excel(writer,'Rev Forecast Committed',header=None,index=False)
    #sheet4
    df4=pass_rev.loc[pass_rev.ClientEx_ProcessVP==key[i]]
    df4.to_excel(writer,'Passthrough Revenue',header=None,index=False)
    #sheet5
    df5=opp_inc.loc[opp_inc.ClientEx_ProcessVP==key[i]]
    df5.to_excel(writer,'Opportunities Included',header=None,index=False)
    #sheet6
    df6=rev_f_fin.loc[rev_f_fin.ClientEx_ProcessVP==key[i]]
    df6.to_excel(writer,'Revenue Forecast Final',header=None,index=False)
    #sheet7
    df7=qoq.loc[qoq.ClientEx_ProcessVP==key[i]]
    df7.to_excel(writer,'QoQ Details',header=None,index=False)
    #sheet8
    df8=mom.loc[mom.ClientEx_ProcessVP==key[i]]
    df8.to_excel(writer,'MoM Details with Location',header=None,index=False)
    #sheet 9,10,11,12,13
    summary.to_excel(writer,'Summary',header=None,index=False)
    hist_data.to_excel(writer,'Historic Data',header=None,index=False)
    raw_data.to_excel(writer,'RawData',header=None,index=False)
    dashboard.to_excel(writer,'Dashboard',header=None,index=False)
    fx.to_excel(writer,'Fx',header=None,index=False)
    #sheet14
    df10=fte18.loc[fte18.ClientEx_ProcessVP==key[i]]
    df10.to_excel(writer,'FTE Forecast- 2018',header=None,index=False)
    #sheet15
    df11=rev18.loc[rev18.ClientEx_ProcessVP==key[i]]
    df11.to_excel(writer,'Rev Forecast- 2018',header=None,index=False)    
    #saving files
    writer.save()

#--------------------------------------------------------------------part2: writing multiheaders and formatting of splitted files-----------------------------------
#6 making list of output files for writting formatting
path = r'C:\Users\pankaj111056\Desktop\FP&A\Automation_FP&A\Version1.0\Test\Splitting'
files = os.listdir(path)
files_xlsx = [f for f in files if f[-4:] == 'xlsx']
#removing master file out of list
files_xlsx.remove('Consolidated Forecast.xlsx')

#7 taking files from the list for formatting
for f in files_xlsx:
    wb = load_workbook(f,data_only=False)
    sheet_list=wb.get_sheet_names()

    #7.1 gettting list of sheets
    modify_raw=wb.get_sheet_by_name(sheet_list[0])#Sheet1- Modify Data
    fte_raw=wb.get_sheet_by_name(sheet_list[1])#Sheet2- FTE Forecast-2017
    rev_f_com_raw=wb.get_sheet_by_name(sheet_list[2])#sheet3- Rev Forecast Committed
    pass_rev_raw=wb.get_sheet_by_name(sheet_list[3])#sheet4- Passthrough Revenue
    opp_inc_raw=wb.get_sheet_by_name(sheet_list[4])#sheet5- Opportunities Included
    rev_f_fin_raw=wb.get_sheet_by_name(sheet_list[5])#sheet6- Revenue Forecast Final
    qoq_raw=wb.get_sheet_by_name(sheet_list[6])#sheet7- QoQ Details
    mom_raw=wb.get_sheet_by_name(sheet_list[7])#sheet8- MoM Details with Location
    summary_raw=wb.get_sheet_by_name(sheet_list[8])#sheet9- Summary
    hist_data=wb.get_sheet_by_name(sheet_list[9])#sheet10- Historic Data
    raw_data=wb.get_sheet_by_name(sheet_list[10])#sheet11- Raw Data
    dashboard=wb.get_sheet_by_name(sheet_list[11])#sheet12- Dashboard
    fx=wb.get_sheet_by_name(sheet_list[12])#sheet13- Fx
    fte18_raw=wb.get_sheet_by_name(sheet_list[13])#sheet14- FTE Forecast- 2018
    rev18_raw=wb.get_sheet_by_name(sheet_list[14])#sheet15- Rev Forecast- 2018
    
    #7.2 changing openpyxl sheets into pandas dataframes which have no headers
    modify = pd.DataFrame(modify_raw.values)
    fte = pd.DataFrame(fte_raw.values)
    rev_f_com = pd.DataFrame(rev_f_com_raw.values)
    pass_rev = pd.DataFrame(pass_rev_raw.values)
    opp_inc = pd.DataFrame(opp_inc_raw.values)
    rev_f_fin = pd.DataFrame(rev_f_fin_raw.values)
    qoq = pd.DataFrame(qoq_raw.values)
    mom = pd.DataFrame(mom_raw.values)
    summary = pd.DataFrame(summary_raw.values)
    hist_data= pd.DataFrame(hist_data.values)
    raw_data= pd.DataFrame(raw_data.values)
    dashboard= pd.DataFrame(dashboard.values)
    fx= pd.DataFrame(fx.values)
    fte18 = pd.DataFrame(fte18_raw.values)
    rev18 = pd.DataFrame(rev18_raw.values)

    #sheet1- Modify Data---------------------------------------------------------------
    #7.3.1 adding headers by concatenating another dataframe with multi-headers
    df1=pd.DataFrame([
        ['','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','',],
        ['Client Code', 'Client Name', 'Client Executive', 'Ops Lead', 'Process VP', 'ClientEx_ProcessVP', 'Vertical',
        'Location','Client Status', 'Modify Data (Yes/No)','Blank Col1','Blank Col2','Blank Col3','Key','Final Key','Sort Key']])
    df=pd.concat([df1,modify])

    #writing dataframes with headers on same file with xlsxwriter
    writer = pd.ExcelWriter(f,engine='xlsxwriter')
    df.to_excel(writer,'Modify Data',header=False,index=False)
    
    #using xlsxwriter for all types of formatting
    #Taking each file as workbook and each of 11 sheets as workbook
    workbook = writer.book
    #7.3.2 defining standard formats which can be used in any xlsxwriter worksheet
    border_frmt=workbook.add_format({'border':1})
    centre_align=workbook.add_format({'align': 'centre','border':1,'num_format': '#,##0;(#,##0);"-"'})
    percent_format = workbook.add_format({'align': 'centre','border':1,'num_format': '0.00%'})
    
    pink_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#e6b8b7','num_format': '#,##0;(#,##0);"-"'})
    light_pink_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#FF99FF','num_format': '#,##0;(#,##0);"-"'})
    dark_pink_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#c6a68c','num_format': '#,##0;(#,##0);"-"'})
    dark_tan_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#c4bd97','num_format': '#,##0;(#,##0);"-"'})
    light_aqua_green_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#b1e9ba','num_format': '#,##0;(#,##0);"-"'})
    dark_aqua_green_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#a4f6cf','num_format': '#,##0;(#,##0);"-"'})
    light_yellow_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#FFFF66','num_format': '#,##0;(#,##0);"-"'})
    dark_yellow_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#c1b407','num_format': '#,##0;(#,##0);"-"'})
    green_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#d3d3d3','num_format': '#,##0;(#,##0);"-"'})
    dark_green_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#2b993d','num_format': '#,##0;(#,##0);"-"'})
    purple_table=workbook.add_format({'align': 'centre','border':1, 'bg_color': '#b1a0c7','num_format': '#,##0;(#,##0);"-"'})
    
    grey_frmt = workbook.add_format({'bg_color': '#d3d3d3',
                                              'border':1})
    light_green_frmt = workbook.add_format({'bg_color': '#98FF98',
                                              'border':1})
    violet_frmt = workbook.add_format({'bg_color': '#b27ccf',
                                              'border':1})
    pink_frmt = workbook.add_format({'bg_color': '#ffc0cb',
                                              'border':1})
    light_red_frmt = workbook.add_format({'bg_color': '#cf7c8f',
                                              'border':1})
    light_blue_frmt = workbook.add_format({'bg_color': '#93bbdb',
                                              'border':1})
    violet_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'bg_color': '#ccc0da'})
    pink_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'bg_color': '#e6b8b7'})
    dark_orange_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'font_color': 'white',
                    'bg_color': '#974706'})
    light_orange_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'bg_color': '#fabf8f'})
    green_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'bg_color': '#b2de82'})
    yellow_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'bg_color': '#FFFF00'})
    light_blue_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'bg_color': '#b7dee8'})
    light_green_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'bg_color': '#d8e4bc'})
    dark_blue_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'font_color': 'white',
                    'bg_color': '#244062'})
    light_green_header_frmt= workbook.add_format({
                    'bold':1,
                    'align': 'left',
                    'border':1,
                    'bg_color': '#d8e4bc'})
    light_blue_header_frmt= workbook.add_format({
                    'bold':1,
                    'align': 'left',
                    'border':1,
                    'bg_color': '#b7dee8'})
    white_header_frmt= workbook.add_format({
                    'bold':1,
                    'bg_color': '#ffffff'})
    sheet_frmt=workbook.add_format({'bg_color': '#ffffff','border':0}) #white color and no border
    sheet_frmt_white_font=workbook.add_format({'bg_color': '#ffffff','border':0, 'font_color': 'white'})
    bold_frmt = workbook.add_format({'bold': 1, 'bg_color': '#FFFF00'})

    #7.3.3 taking each sheet as xlsxwriter worksheet after writing headers
    worksheet1 = writer.sheets['Modify Data']
    
    #column format
    worksheet1.set_column('A:A', 12,border_frmt)
    worksheet1.set_column('B:B', 25,border_frmt)
    worksheet1.set_column('C:E', 15,border_frmt)
    worksheet1.set_column('F:F', 20,border_frmt)
    worksheet1.set_column('G:I', 15,border_frmt)
    worksheet1.set_column('J:J', 20,border_frmt)
    worksheet1.set_column('K:M', 12,border_frmt,{'hidden': True})
    worksheet1.set_column('N:P', 20,centre_align,{'hidden': True})
    #row format
    worksheet1.set_row(0,18,sheet_frmt)
    worksheet1.set_row(1,18,sheet_frmt)
    worksheet1.set_row(2,18,sheet_frmt)
    worksheet1.set_row(3,20,light_green_header_frmt)
    #blanks for columns ahead
    worksheet1.set_column('Q:AZ',12,sheet_frmt)
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet1.set_row(r,12,sheet_frmt)
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet1.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet1.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #setting zoom level
    worksheet1.set_zoom(80)
    
    #sheet2 - FTE Forecast- 2017----------------------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df2=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','',],
        ['Client Code','Client Name','Client Executive', 'Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
        'Client Status', 'Blank Col1','Blank Col2', 'Blank Col3','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17',
        'Dec 17','FY 17','','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','','Jan 17', 'Feb 17',
        'Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
        'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Key','Sort Key']])
    df=pd.concat([df2,fte])
    
    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'FTE Forecast- 2017',header=False,index=False)
    
    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet2 = writer.sheets['FTE Forecast- 2017']
    
    #merging multi-header cells
    worksheet2.merge_range('M3:Y3','COMMITTED SIGNED FTEs', violet_merge_format)
    worksheet2.merge_range('AA3:AM3','TOTAL FTEs (Committed plus Opportunity)', violet_merge_format)
    worksheet2.merge_range('AO3:BA3','TOTAL Headcount', violet_merge_format)
    worksheet2.merge_range('BC3:BO3','MEI',violet_merge_format)
    #column format
    worksheet2.set_column('A:A', 12,border_frmt)
    worksheet2.set_column('B:B', 25,border_frmt)
    worksheet2.set_column('C:E', 15,border_frmt)
    worksheet2.set_column('F:F', 20,border_frmt,{'hidden': True})
    worksheet2.set_column('G:G', 15,border_frmt)
    worksheet2.set_column('H:H', 15,border_frmt)
    worksheet2.set_column('I:I', 15,border_frmt)
    worksheet2.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet2.set_column('M:BO',8,centre_align)
    worksheet2.set_column('BP:BQ',20,centre_align,{'hidden': True})
    #blank columns between tables
    worksheet2.set_column('Z:Z',10,sheet_frmt)
    worksheet2.set_column('AN:AN',10,sheet_frmt)
    worksheet2.set_column('BB:BB',10,sheet_frmt)
    #header row blank cells
    worksheet2.write('Z3','',sheet_frmt)
    worksheet2.write('AN3','',sheet_frmt)
    worksheet2.write('BB3','',sheet_frmt)
    worksheet2.write('Z4','',sheet_frmt)
    worksheet2.write('AN4','',sheet_frmt)
    worksheet2.write('BB4','',sheet_frmt)
    #row format
    worksheet2.set_row(0,18,sheet_frmt)
    worksheet2.set_row(1,18,sheet_frmt)
    worksheet2.set_row(2,20,white_header_frmt)
    worksheet2.set_row(3,20,light_green_header_frmt)
    #freezing
    worksheet2.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet2.set_column('BR:CZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet2.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet2.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet2.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet2.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #zoom level
    worksheet2.set_zoom(80)
    
    #sheet3- Rev Forecast Committed--------------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df3=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical',
        'Location','Client Status','Blank col1','Blank col2','Blank col3','Billing Rate','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17',
        'Sep 17','Oct 17','Nov 17','Dec 17','FY 17','','Q1 17','Q2 17','Q3 17','Q4 17','FY 17',' ', 'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17',
        'Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
        'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total','','Reasons for Variance against Previous forecast',
        'Key','Sort Key']])
    df=pd.concat([df3,rev_f_com])
    
    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'Rev Forecast Committed',header=False,index=False)

    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet3 = writer.sheets['Rev Forecast Committed']

    #merging multi-header cells
    worksheet3.merge_range('N3:AF3','CURRENT FORECAST', light_green_merge_format)
    worksheet3.merge_range('AH3:AZ3','PREVIOUS FORECAST', pink_merge_format)
    worksheet3.merge_range('BB3:BT3','Variance (Current Forecast - Previous Forecast)', light_green_merge_format)
    #column format
    worksheet3.set_column('A:A', 12,border_frmt)
    worksheet3.set_column('B:B', 25,border_frmt)
    worksheet3.set_column('C:E', 15,border_frmt)
    worksheet3.set_column('F:F', 20,border_frmt, {'hidden': True})
    worksheet3.set_column('G:G', 15,border_frmt)
    worksheet3.set_column('H:H', 15,border_frmt)
    worksheet3.set_column('I:I', 15,border_frmt)
    worksheet3.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet3.set_column('M:BT', 15, centre_align)
    worksheet3.set_column('BV:BV', 60,centre_align)
    worksheet3.set_column('BW:BX', 20, centre_align,{'hidden': True})
    #blank columns between tables
    worksheet3.set_column('AA:AA',10,sheet_frmt)
    worksheet3.set_column('AG:AG',10,sheet_frmt)
    worksheet3.set_column('AU:AU',10,sheet_frmt)
    worksheet3.set_column('BA:BA',10,sheet_frmt)
    worksheet3.set_column('BO:BO',10,sheet_frmt)
    worksheet3.set_column('BU:BV',10,sheet_frmt)
    #header row blank cells
    worksheet3.write('AA4','',sheet_frmt)
    worksheet3.write('AG3','',sheet_frmt)
    worksheet3.write('AG4','',sheet_frmt)
    worksheet3.write('AU4','',sheet_frmt)
    worksheet3.write('BA3','',sheet_frmt)
    worksheet3.write('BA4','',sheet_frmt)
    worksheet3.write('BO4','',sheet_frmt)
    worksheet3.write('BU3','',sheet_frmt)
    worksheet3.write('BU4','',sheet_frmt)
    #row format
    worksheet3.set_row(0,18,sheet_frmt)
    worksheet3.set_row(1,18,sheet_frmt)
    worksheet3.set_row(2,20,white_header_frmt)
    worksheet3.set_row(3,20,light_green_header_frmt)
    #freezing
    worksheet3.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet3.set_column('BY:CZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet3.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet3.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet3.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet3.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #zoom level
    worksheet3.set_zoom(80)
    
    #sheet4- Passthrough Revenue--------------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df4=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
         '','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
         '','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
         '','','','','','','','','','','','','','','','','','','',],
        ['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical',
        'Location', 'Client Status', 'Blank col1','Blank col2','Blank col3','Billing Rate','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17',
        'Sep 17','Oct 17','Nov 17','Dec 17','FY 17','','Q1 17','Q2 17','Q3 17','Q4 17','FY 17',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17',
        'Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
        'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17','Total','','Reasons for Variance against Previous forecast','',
        'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17','Q4 17',
        'Total',' ','Jan 17','Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total','','Q1 17','Q2 17','Q3 17',
        'Q4 17','Total','','Reasons for Variance against Budget','Key','Sort Key']])
    df=pd.concat([df4,pass_rev])
    
    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'Passthrough Revenue',header=False,index=False)

    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet4 = writer.sheets['Passthrough Revenue']

    #merging multi-header cells
    worksheet4.merge_range('AH3:AZ3','PREVIOUS FORECAST', pink_merge_format)
    worksheet4.merge_range('BB3:BT3','Variance (Current Forecast - Previous Forecast)', light_green_merge_format)
    worksheet4.merge_range('BX3:CP3','BUDGET', light_blue_merge_format)
    worksheet4.merge_range('CR3:DJ3','Variance (Current Forecast - Budget)', light_green_merge_format)
    #column format
    worksheet4.set_column('A:A', 12,border_frmt)
    worksheet4.set_column('B:B', 25,border_frmt)
    worksheet4.set_column('C:E', 15,border_frmt)
    worksheet4.set_column('F:F', 20,border_frmt, {'hidden': True})
    worksheet4.set_column('G:G', 15,border_frmt)
    worksheet4.set_column('H:H', 15,border_frmt)
    worksheet4.set_column('I:I', 15,border_frmt)
    worksheet4.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet4.set_column('M:BT', 15, centre_align)
    worksheet4.set_column('BV:BV', 45,centre_align)
    worksheet4.set_column('BX:DJ', 15, centre_align)
    worksheet4.set_column('DL:DL', 45,centre_align)
    worksheet4.set_column('DM:DN', 20, centre_align,{'hidden': True})
    #blank columns between tables
    worksheet4.set_column('AA:AA',10,sheet_frmt)
    worksheet4.set_column('AG:AG',10,sheet_frmt)
    worksheet4.set_column('AU:AU',10,sheet_frmt)
    worksheet4.set_column('BA:BA',10,sheet_frmt)
    worksheet4.set_column('BO:BO',10,sheet_frmt)
    worksheet4.set_column('BU:BU',10,sheet_frmt)
    worksheet4.set_column('BW:BW',10,sheet_frmt)
    worksheet4.set_column('CK:CK',10,sheet_frmt)
    worksheet4.set_column('CQ:CQ',10,sheet_frmt)
    worksheet4.set_column('DE:DE',10,sheet_frmt)
    worksheet4.set_column('DK:DK',10,sheet_frmt)
    #header row blank cells
    worksheet4.write('AA4','',sheet_frmt)
    worksheet4.write('AG3','',sheet_frmt)
    worksheet4.write('AG4','',sheet_frmt)
    worksheet4.write('AU4','',sheet_frmt)
    worksheet4.write('BA3','',sheet_frmt)
    worksheet4.write('BA4','',sheet_frmt)
    worksheet4.write('BO4','',sheet_frmt)
    worksheet4.write('BU3','',sheet_frmt)
    worksheet4.write('BU4','',sheet_frmt)
    worksheet4.write('BW3','',sheet_frmt)
    worksheet4.write('BW4','',sheet_frmt)
    worksheet4.write('CK4','',sheet_frmt)
    worksheet4.write('CQ3','',sheet_frmt)
    worksheet4.write('CQ4','',sheet_frmt)
    worksheet4.write('DE4','',sheet_frmt)
    worksheet4.write('DK3','',sheet_frmt)
    worksheet4.write('DK4','',sheet_frmt)
    #row format
    worksheet4.set_row(0,18,sheet_frmt)
    worksheet4.set_row(1,18,sheet_frmt)
    worksheet4.set_row(2,20,white_header_frmt)
    worksheet4.set_row(3,20,light_green_header_frmt)
    #freezing
    worksheet4.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet4.set_column('DO:EZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet4.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet4.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet4.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet4.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #zoom level
    worksheet4.set_zoom(80)
    
    #sheet5- Opportunities Included-----------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df5=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','',],
        ['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
        'Client Status','Blank col1','Blank col2','Blank col3','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17',
        'Dec 17','Total',' ', 'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17', 'Total', ' ','Q1 17',
        'Q2 17','Q3 17','Q4 17','Total', ' ', 'Probability', 'Committed - Unsigned', 'Yet to be won*', 'Total', 'Check', ' ','Jan 17', 'Feb 17','Mar 17','Apr 17',
        'May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total',' ','Q1 17','Q2 17','Q3 17','Q4 17','Total', ' ','Jan 17', 'Feb 17','Mar 17',
        'Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total',' ','Q1 17','Q2 17','Q3 17','Q4 17','Total', ' ','Jan 17', 'Feb 17',
        'Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','Total',' ','Committed Unsigned details', 'Yet-to-be won details',
        'Key','Final Key','Sort Key']])
    df=pd.concat([df5,opp_inc])

    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'Opportunities Included',header=False,index=False)

    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet5 = writer.sheets['Opportunities Included']
    
    #merging multi-header cells
    worksheet5.merge_range('M3:Y3','MoM- Opportunity FTEs (At 100% probablility)', light_blue_merge_format)
    worksheet5.merge_range('AA3:AM3','MoM- Opportunity at 100% Probability (in USD)', light_blue_merge_format)
    worksheet5.merge_range('AO3:AS3','QoQ- Opportunity at 100% Probability', light_blue_merge_format)
    worksheet5.merge_range('AV3:AY3','Total- Probability Based weighted Opportunity', light_blue_merge_format)
    worksheet5.merge_range('BA3:BM3','Probability Weighted Opportunity - MoM (In USD)', light_blue_merge_format)
    worksheet5.merge_range('BO3:BS3','Probability weighted QoQ', light_blue_merge_format)
    worksheet5.merge_range('BU3:CG3','Probability Weighted FTEs - MoM', pink_merge_format)
    worksheet5.merge_range('CI3:CM3','Probability weighted FTEs QoQ', pink_merge_format)
    worksheet5.merge_range('CO3:DA3','2017 Budgeted Opportunity (Weighted)', light_blue_merge_format)
    #column format
    worksheet5.set_column('A:A', 12,border_frmt)
    worksheet5.set_column('B:B', 25,border_frmt)
    worksheet5.set_column('C:E', 15,border_frmt)
    worksheet5.set_column('F:F', 20,border_frmt, {'hidden': True})
    worksheet5.set_column('G:G', 15,border_frmt)
    worksheet5.set_column('H:H', 15,border_frmt)
    worksheet5.set_column('I:I', 15,border_frmt)
    worksheet5.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet5.set_column('M:DA', 12,centre_align)
    worksheet5.set_column('AU:AU', 12,percent_format)
    worksheet5.set_column('AV:DA', 12,centre_align)
    worksheet5.set_column('DC:DD', 60,border_frmt)
    worksheet5.set_column('DE:DG', 20,centre_align,{'hidden': True})
    #blank columns between tables
    worksheet5.set_column('Z:Z',10,sheet_frmt)
    worksheet5.set_column('AN:AN',10,sheet_frmt)
    worksheet5.set_column('AT:AT',10,sheet_frmt)
    worksheet5.set_column('BN:BN',10,sheet_frmt)
    worksheet5.set_column('BT:BT',10,sheet_frmt)
    worksheet5.set_column('CH:CH',10,sheet_frmt)
    worksheet5.set_column('CN:CN',10,sheet_frmt)
    worksheet5.set_column('DB:DB',10,sheet_frmt)
    #header row blank cells
    worksheet5.write('Z3','',sheet_frmt)
    worksheet5.write('Z4','',sheet_frmt)
    worksheet5.write('AN3','',sheet_frmt)
    worksheet5.write('AN4','',sheet_frmt)
    worksheet5.write('AT3','',sheet_frmt)
    worksheet5.write('AT4','',sheet_frmt)
    worksheet5.write('AU3','',sheet_frmt)
    worksheet5.write('AZ3','',sheet_frmt)
    worksheet5.write('AZ4','',sheet_frmt)
    worksheet5.write('BN3','',sheet_frmt)
    worksheet5.write('BN4','',sheet_frmt)
    worksheet5.write('BT3','',sheet_frmt)
    worksheet5.write('BT4','',sheet_frmt)
    worksheet5.write('CH3','',sheet_frmt)
    worksheet5.write('CH4','',sheet_frmt)
    worksheet5.write('CN3','',sheet_frmt)
    worksheet5.write('CN4','',sheet_frmt)
    worksheet5.write('DB3','',sheet_frmt)
    worksheet5.write('DB4','',sheet_frmt)
    worksheet5.write('DC3','',sheet_frmt)
    worksheet5.write('DD3','',sheet_frmt)
    #row format
    worksheet5.set_row(0,18,sheet_frmt)
    worksheet5.set_row(1,18,sheet_frmt)
    worksheet5.set_row(2,20,white_header_frmt)
    worksheet5.set_row(3,20,light_blue_header_frmt)
    #freezing
    worksheet5.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet5.set_column('DH:EZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet5.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet5.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet5.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet5.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #zoom level
    worksheet5.set_zoom(80)
    
    #sheet6- Revenue Forecast Final--------------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df6=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical',
        'Location','Client Status','LoB','Blank col2','Blank col3','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17',
        'Nov 17','Dec 17','Total',' ','Q1 17','Q2 17','Q3 17','Q4 17','Total', ' ','Committed Signed', 'Committed Unsigned', 'Yet to be won*', 'Total', 'Check',
        'Q1 PF', 'var','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17',' ','Q1 17','Q2 17',
        'Q3 17','Q4 17','FY 17',' ', 'Committed Signed', 'Committed Unsigned', 'Yet to be won*', 'FY 17', 'Check',' ', 'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17',
        'Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17',' ','Q1 17','Q2 17','Q3 17','Q4 17','Total',' ','Committed Signed', 'Committed Unsigned',
        'Yet to be won*', 'FY 17',' ','COMMENTS ON VARIANCE AGAINST BUDGET', 'Inputs received','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17',
        'Sep 17','Oct 17','Nov 17','Dec 17','FY 17',' ','Committed Signed', 'Committed Unsigned', 'Yet to be won*', 'FY 17',' ','Jan 17', 'Feb 17','Mar 17','Apr 17',
        'May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17',' ','Committed Signed', 'Committed Unsigned', 'Yet to be won*', 'FY 17',' ',
        'COMMENTS ON VARIANCE(CF-PF)', 'Q1 17 variance', 'Q2 17 variance', 'Q3 17 variance', 'Q4 17 variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
        'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17',
        'Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17',
        'FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ',
        'Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17',
        'Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17',
        'Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17',
        'Oct 17','Nov 17','Dec 17','FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17',
        'FY 17','Variance',' ','Jan 17', 'Feb 17','Mar 17','Apr 17','May 17','Jun 17','Jul 17','Aug 17','Sep 17','Oct 17','Nov 17','Dec 17','FY 17','Variance','Key'
         ,'Sort Key']])
    df=pd.concat([df6,rev_f_fin])

    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'Revenue Forecast Final',header=False,index=False)
    
    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet6 = writer.sheets['Revenue Forecast Final']
    
    #merging multi-header cells
    worksheet6.merge_range('AN3:AZ3','BUDGET', green_merge_format)
    worksheet6.merge_range('BB3:BF3','BUDGET', green_merge_format)
    worksheet6.merge_range('BH3:BL3','BUDGET', green_merge_format)
    worksheet6.merge_range('BN3:BZ3','Variance Against Budget', yellow_merge_format)
    worksheet6.merge_range('CB3:CF3','Variance Against Budget', yellow_merge_format)
    worksheet6.merge_range('CH3:CK3','Variance Against Budget', yellow_merge_format)
    worksheet6.merge_range('CO3:DA3','Previous Forecast', dark_orange_merge_format)
    worksheet6.merge_range('DC3:DF3','Previous Forecast', dark_orange_merge_format)
    worksheet6.merge_range('DH3:DT3','Variance (CF-PF)', dark_blue_merge_format)
    worksheet6.merge_range('DV3:DY3','Variance (CF-PF)', dark_blue_merge_format)
    
    #column format
    worksheet6.set_column('A:A', 12,border_frmt)
    worksheet6.set_column('B:B', 25,border_frmt)
    worksheet6.set_column('C:E', 15,border_frmt)
    worksheet6.set_column('F:F', 20,border_frmt, {'hidden': True})
    worksheet6.set_column('G:G', 15,border_frmt)
    worksheet6.set_column('H:H', 15,border_frmt)
    worksheet6.set_column('I:I', 15,border_frmt)
    worksheet6.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet6.set_column('M:AK', 12,centre_align)
    worksheet6.set_column('AN:AZ', 12,centre_align,{'level': 1})
    worksheet6.set_column('BB:BF', 12,centre_align,{'level': 1})
    worksheet6.set_column('BH:BL', 12,centre_align,{'level': 1})
    worksheet6.set_column('BN:BZ', 12,centre_align,{'level': 1})
    worksheet6.set_column('CA:CA', 12,centre_align)
    worksheet6.set_column('CB:CF', 12,centre_align,{'level': 1})
    worksheet6.set_column('BA:CK', 12,centre_align)
    worksheet6.set_column('CM:CM', 50,centre_align)
    worksheet6.set_column('CN:DY', 12,centre_align)
    worksheet6.set_column('EA:EA', 50,centre_align)
    worksheet6.set_column('EB:EF', 12,centre_align)
    worksheet6.set_column('EG:ET', 12,dark_tan_table,{'level': 1})
    worksheet6.set_column('EV:FI', 12,pink_table,{'level': 1})
    worksheet6.set_column('FK:FX', 12,light_aqua_green_table,{'level': 1})
    worksheet6.set_column('FZ:GM', 12,light_yellow_table,{'level': 1})
    worksheet6.set_column('GO:HB', 12,dark_yellow_table,{'level': 1})
    worksheet6.set_column('HD:HQ', 12,light_pink_table,{'level': 1})
    worksheet6.set_column('HS:IF', 12,dark_aqua_green_table,{'level': 1})
    worksheet6.set_column('IH:IU', 12,dark_green_table,{'level': 1})
    worksheet6.set_column('IW:JJ', 12,dark_pink_table,{'level': 1})    
    worksheet6.set_column('JL:JY', 12,purple_table,{'level': 1})
    worksheet6.set_column('JZ:KA', 12,centre_align,{'hidden': True})
    
    #blank columns between tables
    worksheet6.set_column('Z:Z',10,sheet_frmt_white_font)#---formulas present but text color white
    worksheet6.set_column('AF:AF',10,sheet_frmt_white_font)
    worksheet6.set_column('AL:AM', 5,sheet_frmt_white_font)
    worksheet6.set_column('BA:BA',10,sheet_frmt_white_font)
    worksheet6.set_column('BG:BG',10,sheet_frmt_white_font)
    worksheet6.set_column('BM:BM',10,sheet_frmt_white_font)
    worksheet6.set_column('CA:CA',10,sheet_frmt_white_font)
    worksheet6.set_column('CG:CG',10,sheet_frmt_white_font)
    worksheet6.set_column('CL:CL',10,sheet_frmt_white_font)#----till here formulas with white text
    worksheet6.set_column('DB:DB',10,sheet_frmt)
    worksheet6.set_column('DG:DG',10,sheet_frmt)
    worksheet6.set_column('DU:DU',10,sheet_frmt)
    worksheet6.set_column('DZ:DZ',10,sheet_frmt)
    worksheet6.set_column('EF:EF',10,sheet_frmt)
    worksheet6.set_column('EU:EU',10,sheet_frmt)
    worksheet6.set_column('FJ:FJ',10,sheet_frmt)
    worksheet6.set_column('FY:FY',10,sheet_frmt)
    worksheet6.set_column('GN:GN',10,sheet_frmt)
    worksheet6.set_column('HC:HC',10,sheet_frmt)
    worksheet6.set_column('HR:HR',10,sheet_frmt)
    worksheet6.set_column('IG:IG',10,sheet_frmt)
    worksheet6.set_column('IV:IV',10,sheet_frmt)
    worksheet6.set_column('JK:JK',10,sheet_frmt)
    
    #header row blank cells
    worksheet6.write('Z4','',sheet_frmt)
    worksheet6.write('AF4','',sheet_frmt)
    worksheet6.write('AL3','',sheet_frmt)
    worksheet6.write('AL4','',sheet_frmt)
    worksheet6.write('AM3','',sheet_frmt)
    worksheet6.write('AM4','',sheet_frmt)
    worksheet6.write('BA3','',sheet_frmt)
    worksheet6.write('BA4','',sheet_frmt)
    worksheet6.write('BG3','',sheet_frmt)
    worksheet6.write('BG4','',sheet_frmt)
    worksheet6.write('BM3','',sheet_frmt)
    worksheet6.write('BM4','',sheet_frmt)
    worksheet6.write('CA3','',sheet_frmt)
    worksheet6.write('CA4','',sheet_frmt)
    worksheet6.write('CG3','',sheet_frmt)
    worksheet6.write('CG4','',sheet_frmt)
    worksheet6.write('CL3','',sheet_frmt)
    worksheet6.write('CL4','',sheet_frmt)
    worksheet6.write('DB3','',sheet_frmt)
    worksheet6.write('DB4','',sheet_frmt)
    worksheet6.write('DG3','',sheet_frmt)
    worksheet6.write('DG4','',sheet_frmt)
    worksheet6.write('DU4','',sheet_frmt)
    worksheet6.write('DZ4','',sheet_frmt)
    worksheet6.write('EF4','',sheet_frmt)
    worksheet6.write('EU4','',sheet_frmt)
    worksheet6.write('FJ4','',sheet_frmt)
    worksheet6.write('FY4','',sheet_frmt)
    worksheet6.write('GN4','',sheet_frmt)
    worksheet6.write('HC4','',sheet_frmt)
    worksheet6.write('HR4','',sheet_frmt)
    worksheet6.write('IG4','',sheet_frmt)
    worksheet6.write('IV4','',sheet_frmt)
    worksheet6.write('JK4','',sheet_frmt)
    
    #row format
    worksheet6.set_row(0,18,sheet_frmt)
    worksheet6.set_row(1,18,sheet_frmt)
    worksheet6.set_row(2,20,white_header_frmt)
    worksheet6.set_row(3,20,light_green_header_frmt)
    #freezing
    worksheet6.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet6.set_column('KB:KZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet6.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet6.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet6.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet6.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #zoom level
    worksheet6.set_zoom(80)
    
    #sheet7- QoQ Details---------------------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df7=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location','Client Status',
        'LoB','Blank col2','Blank col3','Committed Signed', 'Committed - Unsigned','Yet to be won*','NCR','TOTAL','','Committed Signed',
        'Committed - Unsigned','Yet to be won*','NCR','TOTAL','','Committed Signed', 'Committed - Unsigned','Yet to be won*','NCR','TOTAL','',
        'Committed Signed', 'Committed - Unsigned','Yet to be won*','NCR','TOTAL','','Committed Signed', 'Committed - Unsigned','Yet to be won*','NCR','TOTAL',
        'Key','Final Key','Sort Key']])
    df=pd.concat([df7,qoq])
    
    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'QoQ Details',header=False,index=False)
    
    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet7 = writer.sheets['QoQ Details']
    
    #merging multi-header cells
    worksheet7.merge_range('M3:Q3','Q1 17', green_merge_format)
    worksheet7.merge_range('S3:W3','Q2 17', green_merge_format)
    worksheet7.merge_range('Y3:AC3','Q3 17', green_merge_format)
    worksheet7.merge_range('AE3:AI3','Q4 17', green_merge_format)
    worksheet7.merge_range('AK3:AO3','FY 17', green_merge_format)
    #column format
    worksheet7.set_column('A:A', 12,border_frmt)
    worksheet7.set_column('B:B', 25,border_frmt)
    worksheet7.set_column('C:E', 15,border_frmt)
    worksheet7.set_column('F:F', 20,border_frmt, {'hidden': True})
    worksheet7.set_column('G:G', 15,border_frmt)
    worksheet7.set_column('H:H', 15,border_frmt)
    worksheet7.set_column('I:I', 15,border_frmt)
    worksheet7.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet7.set_column('M:AO',12,centre_align)
    worksheet7.set_column('AP:AP',20,centre_align,{'hidden': True})
    worksheet7.set_column('AQ:AR',10,centre_align,{'hidden': True})
    #blank columns between tables
    worksheet7.set_column('R:R',10,sheet_frmt)
    worksheet7.set_column('X:X',10,sheet_frmt)
    worksheet7.set_column('AD:AD',10,sheet_frmt)
    worksheet7.set_column('AJ:AJ',10,sheet_frmt)
    #header row blank cells
    worksheet7.write('R3','',sheet_frmt)
    worksheet7.write('R4','',sheet_frmt)
    worksheet7.write('X3','',sheet_frmt)
    worksheet7.write('X4','',sheet_frmt)
    worksheet7.write('AD3','',sheet_frmt)
    worksheet7.write('AD4','',sheet_frmt)
    worksheet7.write('AJ3','',sheet_frmt)
    worksheet7.write('AJ4','',sheet_frmt)
    #row format
    worksheet7.set_row(0,18,sheet_frmt)
    worksheet7.set_row(1,18,sheet_frmt)
    worksheet7.set_row(2,20,white_header_frmt)
    worksheet7.set_row(3,20,light_green_header_frmt)
    #freezing
    worksheet7.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet7.set_column('AS:BZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet7.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet7.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet7.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations- not required

    #zoom level
    worksheet7.set_zoom(80)
    
    #sheet8- MoM Details with Location-----------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df8=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','',],
        ['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
        'Client Status','LoB','Blank col2','Blank col3','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed',
        'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
        'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
        'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
        'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
        'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','','Committed Signed', 'Committed - Unsigned',
        'Yet to be won*','','Committed Signed', 'Committed - Unsigned','Yet to be won*','Key','Sort Key']])
    df=pd.concat([df8,mom])
    
    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'MoM Details with Location',header=False,index=False)
    
    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet8 = writer.sheets['MoM Details with Location']
    
    #merging multi-header cells
    worksheet8.merge_range('M3:O3','Jan 17', green_merge_format)
    worksheet8.merge_range('Q3:S3','Feb 17', green_merge_format)
    worksheet8.merge_range('U3:W3','Mar 17', green_merge_format)
    worksheet8.merge_range('Y3:AA3','Apr 17', green_merge_format)
    worksheet8.merge_range('AC3:AE3','May 17', green_merge_format)
    worksheet8.merge_range('AG3:AI3','Jun 17', green_merge_format)
    worksheet8.merge_range('AK3:AM3','Jul 17', green_merge_format)
    worksheet8.merge_range('AO3:AQ3','Aug 17', green_merge_format)
    worksheet8.merge_range('AS3:AU3','Sep 17', green_merge_format)
    worksheet8.merge_range('AW3:AY3','Oct 17', green_merge_format)
    worksheet8.merge_range('BA3:BC3','Nov 17', green_merge_format)
    worksheet8.merge_range('BE3:BG3','Dec 17', green_merge_format)
    worksheet8.merge_range('BI3:BK3','FY 17', green_merge_format)
    #column format
    worksheet8.set_column('A:A', 12,border_frmt)
    worksheet8.set_column('B:B', 25,border_frmt)
    worksheet8.set_column('C:E', 15,border_frmt)
    worksheet8.set_column('F:F', 20,border_frmt, {'hidden': True})
    worksheet8.set_column('G:G', 15,border_frmt)
    worksheet8.set_column('H:H', 15,border_frmt)
    worksheet8.set_column('I:I', 15,border_frmt)
    worksheet8.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet8.set_column('M:BK',15,centre_align)
    worksheet8.set_column('BL:BM',20,centre_align,{'hidden': True})
    #blank columns between tables
    worksheet8.set_column('P:P',10,sheet_frmt)
    worksheet8.set_column('T:T',10,sheet_frmt)
    worksheet8.set_column('X:X',10,sheet_frmt)
    worksheet8.set_column('AB:AB',10,sheet_frmt)
    worksheet8.set_column('AF:AF',10,sheet_frmt)
    worksheet8.set_column('AJ:AJ',10,sheet_frmt)
    worksheet8.set_column('AN:AN',10,sheet_frmt)
    worksheet8.set_column('AV:AV',10,sheet_frmt)
    worksheet8.set_column('AZ:AZ',10,sheet_frmt)
    worksheet8.set_column('BD:BD',10,sheet_frmt)
    worksheet8.set_column('BH:BH',10,sheet_frmt)
    #header row blank cells
    worksheet8.write('P3','',sheet_frmt)
    worksheet8.write('P4','',sheet_frmt)
    worksheet8.write('T3','',sheet_frmt)
    worksheet8.write('T4','',sheet_frmt)
    worksheet8.write('X3','',sheet_frmt)
    worksheet8.write('X4','',sheet_frmt)
    worksheet8.write('AB3','',sheet_frmt)
    worksheet8.write('AB4','',sheet_frmt)
    worksheet8.write('AF3','',sheet_frmt)
    worksheet8.write('AF4','',sheet_frmt)
    worksheet8.write('AJ3','',sheet_frmt)
    worksheet8.write('AJ4','',sheet_frmt)
    worksheet8.write('AN3','',sheet_frmt)
    worksheet8.write('AN4','',sheet_frmt)
    worksheet8.write('AV3','',sheet_frmt)
    worksheet8.write('AV4','',sheet_frmt)
    worksheet8.write('AZ3','',sheet_frmt)
    worksheet8.write('AZ4','',sheet_frmt)
    worksheet8.write('BD3','',sheet_frmt)
    worksheet8.write('BD4','',sheet_frmt)
    worksheet8.write('BH3','',sheet_frmt)
    worksheet8.write('BH4','',sheet_frmt)
    #row format
    worksheet8.set_row(0,18,sheet_frmt)
    worksheet8.set_row(1,18,sheet_frmt)
    worksheet8.set_row(2,20,white_header_frmt)
    worksheet8.set_row(3,20,light_green_header_frmt)
    #freezing
    worksheet8.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet8.set_column('BN:CZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet8.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet8.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet8.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet8.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #zoom level
    worksheet8.set_zoom(80)

    #sheet 9,10,11,12,13- Summary, Historic Data, RawData, Dashboard, Fx----------------------------------------------------------------
    #No formatting required as the sheet is hidden in split files
    summary.to_excel(writer,'Summary',header=False,index=False)
    worksheet9 = writer.sheets['Summary']
    worksheet9.hide()

    hist_data.to_excel(writer,'Historic Data',header=False,index=False)
    worksheet_hist_data = writer.sheets['Historic Data']
    worksheet_hist_data.hide()
    
    raw_data.to_excel(writer,'RawData',header=False,index=False)
    worksheet_raw_data = writer.sheets['RawData']
    worksheet_raw_data.hide()

    dashboard.to_excel(writer,'Dashboard',header=False,index=False)
    worksheet_dashboard = writer.sheets['Dashboard']
    worksheet_dashboard.hide()

    fx.to_excel(writer,'Fx',header=False,index=False)
    worksheet_fx = writer.sheets['Fx']
    worksheet_fx.hide()
    
    #sheet14- FTE Forecast- 2018------------------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df10=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
        'Client Status','Blank col1','Blank col2','Blank col3','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18',
        'Dec 18','FY 18','','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','',
        'Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','',
        'Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','',
        'Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','Key','Sort Key']])
    df=pd.concat([df10,fte18])
    
    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'FTE Forecast- 2018',header=False,index=False)
    
    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet10 = writer.sheets['FTE Forecast- 2018']
    
    #merging multi-header cells
    worksheet10.merge_range('M3:Y3','COMMITTED SIGNED FTEs', light_blue_merge_format)
    worksheet10.merge_range('AA3:AM3','Weighted opportunity FTEs', light_blue_merge_format)
    worksheet10.merge_range('AO3:BA3','TOTAL Headcount', light_blue_merge_format)
    worksheet10.merge_range('BC3:BO3','Total FTE', light_blue_merge_format)
    worksheet10.merge_range('BQ3:CC3','MEI', light_blue_merge_format)
    #column format
    worksheet10.set_column('A:A', 12,border_frmt)
    worksheet10.set_column('B:B', 25,border_frmt)
    worksheet10.set_column('C:E', 15,border_frmt)
    worksheet10.set_column('F:F', 20,border_frmt, {'hidden': True})
    worksheet10.set_column('G:G', 15,border_frmt)
    worksheet10.set_column('H:H', 15,border_frmt)
    worksheet10.set_column('I:I', 15,border_frmt)
    worksheet10.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet10.set_column('M:CC',12,centre_align)
    worksheet10.set_column('CD:CE',20,centre_align,{'hidden': True})
    #blank columns between tables
    worksheet10.set_column('Z:Z',10,sheet_frmt)
    worksheet10.set_column('AN:AN',10,sheet_frmt)
    worksheet10.set_column('BB:BB',10,sheet_frmt)
    worksheet10.set_column('BP:BP',10,sheet_frmt)
    #header row blank cells
    worksheet10.write('Z3','',sheet_frmt)
    worksheet10.write('Z4','',sheet_frmt)
    worksheet10.write('AN3','',sheet_frmt)
    worksheet10.write('AN4','',sheet_frmt)
    worksheet10.write('BB3','',sheet_frmt)
    worksheet10.write('BB4','',sheet_frmt)
    worksheet10.write('BP3','',sheet_frmt)
    worksheet10.write('BP4','',sheet_frmt)
    #row format
    worksheet10.set_row(0,18,sheet_frmt)
    worksheet10.set_row(1,18,sheet_frmt)
    worksheet10.set_row(2,20,white_header_frmt)
    worksheet10.set_row(3,20,light_green_header_frmt)
    #freezing
    worksheet10.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet10.set_column('CF:DZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet10.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet10.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet10.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet10.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #zoom level
    worksheet10.set_zoom(80)

    #sheet15- Rev Forecast- 2018------------------------------------------------------
    #adding headers by concatenating another dataframe with multi-headers
    df11=pd.DataFrame([
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
        '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',],
        ['Client Code','Client Name','Client Executive','Ops Lead','Process VP','ClientEx_ProcessVP','Vertical','Location',
        'Client Status','Blank col1','Blank col2','Blank col3','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18',
        'Dec 18','FY 18','','Q1 18','Q2 18','Q3 18','Q4 18','FY 18',' ','Probability %','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18','Jun 18','Jul 18','Aug 18',
        'Sep 18','Oct 18','Nov 18','Dec 18','FY 18','','Q1 18','Q2 18','Q3 18','Q4 18','FY 18',' ','Probability %','Jan 18', 'Feb 18','Mar 18','Apr 18','May 18',
        'Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','','Q1 18','Q2 18','Q3 18','Q4 18','FY 18',' ','Jan 18', 'Feb 18','Mar 18','Apr 18',
        'May 18','Jun 18','Jul 18','Aug 18','Sep 18','Oct 18','Nov 18','Dec 18','FY 18','','Q1 18','Q2 18','Q3 18','Q4 18','FY 18',' ','Comments','Key','Sort Key']])
    df=pd.concat([df11,rev18])
    
    #writing dataframes with headers on same file with xlsxwriter
    df.to_excel(writer,'Rev Forecast- 2018',header=False,index=False)
    
    #taking each sheet as xlsxwriter worksheet after writing headers
    worksheet11 = writer.sheets['Rev Forecast- 2018']
    
    #merging multi-header cells
    worksheet11.merge_range('M3:AE3','Committed Revenue Forecast- 2018', light_orange_merge_format)
    worksheet11.merge_range('AH3:AZ3','2018 Committed Unsigned Revenue', pink_merge_format)
    worksheet11.merge_range('BC3:BU3','2018 WEIGHTED OPPORTUNITY REVENUE', pink_merge_format)
    worksheet11.merge_range('BW3:CO3','TOTAL FORECAST REVENUE- 2018', light_green_merge_format)
    #column format
    worksheet11.set_column('A:A', 12,border_frmt)
    worksheet11.set_column('B:B', 25,border_frmt)
    worksheet11.set_column('C:E', 15,border_frmt)
    worksheet11.set_column('F:F', 20,border_frmt, {'hidden': True})
    worksheet11.set_column('G:G', 15,border_frmt)
    worksheet11.set_column('H:H', 15,border_frmt)
    worksheet11.set_column('I:I', 15,border_frmt)
    worksheet11.set_column('J:L', 15,border_frmt,{'hidden': True})
    worksheet11.set_column('M:AE',12,centre_align)
    worksheet11.set_column('AG:AG',12,percent_format)
    worksheet11.set_column('AH:BA',12,centre_align)
    worksheet11.set_column('BB:BB',12,percent_format)
    worksheet11.set_column('BC:CO',12,centre_align)
    worksheet11.set_column('CQ:CQ',45,centre_align)
    worksheet11.set_column('CR:CS',20,centre_align,{'hidden': True})
    #blank columns between tables
    worksheet11.set_column('Z:Z',10,sheet_frmt)
    worksheet11.set_column('AF:AF',10,sheet_frmt)
    worksheet11.set_column('AU:AU',10,sheet_frmt)
    worksheet11.set_column('BA:BA',10,sheet_frmt)
    worksheet11.set_column('BV:BV',10,sheet_frmt)
    worksheet11.set_column('BP:BP',10,sheet_frmt)
    worksheet11.set_column('CJ:CJ',10,sheet_frmt)
    worksheet11.set_column('CP:CP',10,sheet_frmt)
    #header row blank cells
    worksheet11.write('Z4','',sheet_frmt)
    worksheet11.write('AF3','',sheet_frmt)
    worksheet11.write('AF4','',sheet_frmt)
    worksheet11.write('AU4','',sheet_frmt)
    worksheet11.write('BA3','',sheet_frmt)
    worksheet11.write('BA4','',sheet_frmt)
    worksheet11.write('BP4','',sheet_frmt)
    worksheet11.write('BV3','',sheet_frmt)
    worksheet11.write('BV4','',sheet_frmt)
    worksheet11.write('CJ4','',sheet_frmt)
    worksheet11.write('CP3','',sheet_frmt)
    worksheet11.write('CP4','',sheet_frmt)
    #row format
    worksheet11.set_row(0,18,sheet_frmt)
    worksheet11.set_row(1,18,sheet_frmt)
    worksheet11.set_row(2,20,white_header_frmt)
    worksheet11.set_row(3,20,light_green_header_frmt)
    #freezing
    worksheet11.freeze_panes(4, 8)
    #blanks for columns ahead
    worksheet11.set_column('CT:DZ',12,sheet_frmt) 
    row=len(df.index)
    clm=len(df.columns)
    #set next 50 rows below blank
    for r in range(row,row+50):
        worksheet11.set_row(r,12,sheet_frmt) 
    #set header rows blank
    for c in range(clm, clm+50):
        cell = xl_rowcol_to_cell(3, c)
        worksheet11.write(cell,'',sheet_frmt) 
        cell = xl_rowcol_to_cell(2, c)
        worksheet11.write(cell,'',sheet_frmt)
    #highlighting TOTAL in locations
    worksheet11.conditional_format('H1:H1000', {'type':     'text',
                                       'criteria': 'containing',
                                       'value':    'TOTAL',
                                       'format':   bold_frmt})
    #zoom level
    worksheet11.set_zoom(80)
writer.save()
