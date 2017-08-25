import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell

# define the location of the split files
path = r'\\172.16.2.131\Budget\Budgets\Budget 2018\F&A\Revenue Budget\Received input files\Consolidation'
files = os.listdir(path)
files_xlsx = [f for f in files if f[-4:] == 'xlsx']

trans_data_raw=pd.DataFrame()
for f in files_xlsx:
    wb = load_workbook(f,data_only=True)
    trans_data=wb.get_sheet_by_name('Transformed Data')

    df = pd.DataFrame(trans_data.values, columns=['FPnA Portal Names','Vertical','Process Names', 'Process Code', 'SOW Name/No.', 'SOW Start date', 'SOW End date', 'Deal Classfication',
                                                  'Market served by Client', 'Probability', 'Pricing Type', 'Sub Location', 'Location', 'Location Code',
                                                  'Currency', 'COLA%/(DISCOUNT%)\n 2018', 'COLA%/(DISCOUNT%)\n 2019', 'COLA%/(DISCOUNT%)\n 2020',
                                                  'COLA%/(DISCOUNT%)\n 2021', 'Applicable Month', 'FX Applicability', 'Fx Currency', 'Fx Lower Band',
                                                  'Fx Higher Band', "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18",
                                                  "Oct'18", "Nov'18", "Dec'18", "Q1 '18", "Q2 '18", "Q3 '18", "Q4 '18", "FY '18", "FY'19", "FY'20", "FY'21",
                                                  None, "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18",
                                                  "Nov'18", "Dec'18", "FY'18", "FY'19", "FY'20", "FY'21", None, "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18",
                                                  "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18", "Dec'18", "FY '18", "FY'19", "FY'20", "FY'21", None,
                                                  "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18",
                                                  "Dec'18", "Q1 '18", "Q2 '18", "Q3 '18", "Q4 '18", 'FY-18', "FY'19", "FY'20", "FY'21", None, "Jan'18", "Feb'18",
                                                  "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18", "Dec'18", "Q1 '18",
                                                  "Q2 '18", "Q3 '18", "Q4 '18", "FY '18", "FY'19", "FY'20", "FY'21","Key"])
    df=df[2:]
    trans_data_raw=trans_data_raw.append(df)
trans_data_raw.index = range(len(trans_data_raw))

#-----------------------------------------------------------------part2: selecting the data based on the Yes/No columns-------------------------------------

row_index=np.where(trans_data_raw['Key']=="Yes") [0]
trans_data=pd.DataFrame()
for row in row_index:
        trans_data=trans_data.append(trans_data_raw.loc[[row]])

writer = pd.ExcelWriter('consolidated_budget.xlsx')
trans_data.to_excel(writer,'Transformed Data',header=None,index=False)
writer.save()

#---------------------------------------------------------------part3: writing multiheaders and formatting of the consolidated file--------------------------------
#taking file
file = r'\\172.16.2.131\Budget\Budgets\Budget 2018\F&A\Revenue Budget\Received input files\Consolidation\consolidated_budget.xlsx'
wb = load_workbook(file,data_only=True)
trans_data_sheet=wb.get_sheet_by_name('Transformed Data')
trans_data_df = pd.DataFrame(trans_data_sheet.values)
#adding headers by concatenating another dataframe with multi-headers
df1=pd.DataFrame([
    ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
     '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
     '','','','','','','','','','','','','','','','','','','','','','','','','',],
    ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
     '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
     '','','','','','','','','','','','','','','','','','','','','','','','','',],
    ['','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
     '','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','',
     '','','','','','','','','','','','','','','','','','','','','','','','','',],
    ['FPnA Portal Names','Vertical','Process Names', 'Process Code', 'SOW Name/No.', 'SOW Start date', 'SOW End date', 'Deal Classfication',
                                                  'Market served by Client', 'Probability', 'Pricing Type', 'Sub Location', 'Location', 'Location Code',
                                                  'Currency', 'COLA%/(DISCOUNT%)\n 2018', 'COLA%/(DISCOUNT%)\n 2019', 'COLA%/(DISCOUNT%)\n 2020',
                                                  'COLA%/(DISCOUNT%)\n 2021', 'Applicable Month', 'FX Applicability', 'Fx Currency', 'Fx Lower Band',
                                                  'Fx Higher Band', "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18",
                                                  "Oct'18", "Nov'18", "Dec'18", "Q1 '18", "Q2 '18", "Q3 '18", "Q4 '18", "FY '18", "FY'19", "FY'20", "FY'21",
                                                  None, "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18",
                                                  "Nov'18", "Dec'18", "FY'18", "FY'19", "FY'20", "FY'21", None, "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18",
                                                  "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18", "Dec'18", "FY '18", "FY'19", "FY'20", "FY'21", None,
                                                  "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18",
                                                  "Dec'18", "Q1 '18", "Q2 '18", "Q3 '18", "Q4 '18", 'FY-18', "FY'19", "FY'20", "FY'21", None, "Jan'18", "Feb'18",
                                                  "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18", "Dec'18", "Q1 '18",
                                                  "Q2 '18", "Q3 '18", "Q4 '18", "FY '18", "FY'19", "FY'20", "FY'21","Key"]])
df=pd.concat([df1,trans_data_df])

writer = pd.ExcelWriter(file,engine='xlsxwriter')
df.to_excel(writer,'Transformed Data',header=False,index=False)

#using xlsxwriter for all types of formatting
#Taking each file as workbook and each of 11 sheets as workbook
workbook = writer.book
light_blue_merge_format = workbook.add_format({
                    'bold':1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border':0,
                    'bg_color': '#b7dee8'})
light_blue_header_frmt= workbook.add_format({
                    'bold':1,
                    'align': 'left',
                    'border':1,
                    'bg_color': '#b7dee8'})
no_accounting_frmt=workbook.add_format({'num_format': '0'})
date_frmt=workbook.add_format({'num_format': 'mm/dd/yy'})
percent_frmt = workbook.add_format({'num_format': '0.00%'})
accounting_frmt=workbook.add_format({'num_format': '#,##0;(#,##0);"-"'})
sheet_frmt=workbook.add_format({'bg_color': '#ffffff'})

worksheet = writer.sheets['Transformed Data']

worksheet.merge_range('Y3:AR3','Total Revenue', light_blue_merge_format)
worksheet.merge_range('AT3:BI3','Total Number of Billable FTEs (Probability Weighted)', light_blue_merge_format)
worksheet.merge_range('BK3:BZ3','COLA$/Discount$', light_blue_merge_format)
worksheet.merge_range('CB3:CU3','Total Number of Billable FTEs (Not Probability Weighted)',light_blue_merge_format)
worksheet.merge_range('CW3:DP3','Total Revenue (Not Probability Weighted)',light_blue_merge_format)
#col frmts
worksheet.set_column('A:D', 15,no_accounting_frmt)
worksheet.set_column('F:G', 9,date_frmt)
worksheet.set_column('J:J', 9,percent_frmt)
worksheet.set_column('P:S', 9,percent_frmt)
worksheet.set_column('Y:DP', 9,accounting_frmt)
worksheet.set_column('DQ:DQ', 9,accounting_frmt,{'hidden': True})
#blank cols- white
worksheet.set_column('AS:AS', 9,sheet_frmt)
worksheet.set_column('BJ:BJ', 9,sheet_frmt)
worksheet.set_column('CA:CA', 9,sheet_frmt)
worksheet.set_column('CV:CV', 9,sheet_frmt)

clm=len(df.columns)
clm_list=['FPnA Portal Names','Vertical','Process Names', 'Process Code', 'SOW Name/No.', 'SOW Start date', 'SOW End date', 'Deal Classfication',
                                                  'Market served by Client', 'Probability', 'Pricing Type', 'Sub Location', 'Location', 'Location Code',
                                                  'Currency', 'COLA%/(DISCOUNT%)\n 2018', 'COLA%/(DISCOUNT%)\n 2019', 'COLA%/(DISCOUNT%)\n 2020',
                                                  'COLA%/(DISCOUNT%)\n 2021', 'Applicable Month', 'FX Applicability', 'Fx Currency', 'Fx Lower Band',
                                                  'Fx Higher Band', "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18",
                                                  "Oct'18", "Nov'18", "Dec'18", "Q1 '18", "Q2 '18", "Q3 '18", "Q4 '18", "FY '18", "FY'19", "FY'20", "FY'21",
                                                  None, "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18",
                                                  "Nov'18", "Dec'18", "FY'18", "FY'19", "FY'20", "FY'21", None, "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18",
                                                  "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18", "Dec'18", "FY '18", "FY'19", "FY'20", "FY'21", None,
                                                  "Jan'18", "Feb'18", "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18",
                                                  "Dec'18", "Q1 '18", "Q2 '18", "Q3 '18", "Q4 '18", 'FY-18', "FY'19", "FY'20", "FY'21", None, "Jan'18", "Feb'18",
                                                  "Mar'18", "Apr'18", "May'18", "Jun'18", "Jul'18", "Aug'18", "Sep'18", "Oct'18", "Nov'18", "Dec'18", "Q1 '18",
                                                  "Q2 '18", "Q3 '18", "Q4 '18", "FY '18", "FY'19", "FY'20", "FY'21","Key"] 
    #set header rows blank
for c in range(clm):
    cell = xl_rowcol_to_cell(3, c)
    worksheet.write(cell,clm_list[c],light_blue_header_frmt)
worksheet.freeze_panes(4, 4)
worksheet.set_zoom(80)
writer.save()





