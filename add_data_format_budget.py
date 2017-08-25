import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell

temp_file = r'\\172.16.2.131\Budget\Budgets\Budget 2018\F&A\Revenue Budget\Received input files\data_format.xlsx'
temp = load_workbook(temp_file,data_only=False)
new_sheet=temp.get_sheet_by_name('Transformed Data')
new_data=pd.DataFrame(new_sheet.values)
    
# define the location of the split files
path = r'\\172.16.2.131\Budget\Budgets\Budget 2018\F&A\Revenue Budget\Received input files'
files = os.listdir(path)
files_xlsx = [f for f in files if f[-4:] == 'xlsx']
files_xlsx.remove('data_format.xlsx')

for f in files_xlsx:
    wb = load_workbook(f,data_only=True)
    sheet_list=wb.get_sheet_names()
    #gettting list of sheets

    fte_raw=wb.get_sheet_by_name(sheet_list[3])#sheet4
    bill_raw=wb.get_sheet_by_name(sheet_list[4])#sheet5
    outp_raw=wb.get_sheet_by_name(sheet_list[7])#sheet6
    cola_raw=wb.get_sheet_by_name(sheet_list[6])#sheet8
    
#changing openpyxl sheets into pandas dataframes which have no headers
    fte = pd.DataFrame(fte_raw.values)
    bill = pd.DataFrame(bill_raw.values)
    outp = pd.DataFrame(outp_raw.values)
    cola = pd.DataFrame(cola_raw.values)

    f_new='_'+f
    writer = pd.ExcelWriter(f_new)

#sheet4
    fte.to_excel(writer,'FTE',header=None,index=False)
#sheet5
    bill.to_excel(writer,'Billing Rates',header=None,index=False)
#sheet6
    outp.to_excel(writer,'Output ',header=None,index=False)
#sheet8
    cola.to_excel(writer,'COLA Working',header=None,index=False)
#sheet11
    new_data.to_excel(writer,'Transformed Data',header=None,index=False)
    writer.save()
